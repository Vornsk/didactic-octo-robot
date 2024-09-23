from flask import Flask, render_template, request, jsonify, redirect, url_for, session, send_file
import hashlib, os, json
from weather import weather_list
from account import users
from openpyxl import Workbook
from apscheduler.schedulers.background import BackgroundScheduler
import smtplib
from email.mime.text import MIMEText
from datetime import datetime, timedelta


app = Flask(__name__)

# 할 일 저장 파일
TASKS_FILE = 'tasks.json'
app.secret_key = os.urandom(24)
today = datetime.now().strftime("%Y-%m-%d")
weather_data = weather_list

# 메인 페이지
@app.route('/')
def index():
    username = session.get('username')
    team = session.get('team', '')
    nickname = session.get('nickname', '')
    return render_template('index.html', username=username, team=team, nickname=nickname)

# 로그인 페이지
@app.route('/login')
def login_page():
    return render_template('login.html')

# 로그인 처리
@app.route('/login', methods=['POST'])
def login():
    error = None
    username = request.form.get('username')
    password = request.form.get('password')

    # 비밀번호 해시 처리 (SHA-256)
    hashed_password = hashlib.sha256(password.encode()).hexdigest()

    # 사용자 데이터 가져오기
    user_data = users.get(username)

    # 로그인
    if user_data and user_data['password'] == hashed_password:
        session['username'] = username
        session['team'] = user_data['team']
        session['nickname'] = user_data['nickname']
        return redirect(url_for('index'))
    else:
        error = "로그인 실패"
        return render_template('login.html', error=error)

# 로그아웃 처리
@app.route('/logout')
def logout():
    session.pop('username', None)
    session.pop('team', None)
    session.pop('nickname', None)
    return redirect(url_for('index'))

@app.route('/calendar')
def go_to_calendar():
    username = session.get('username')
    team = session.get('team', '')
    nickname = session.get('nickname', '')
    if not username:
        return redirect(url_for('login'))
    return render_template('calendar.html', username=username, team=team, nickname=nickname)


def load_tasks():
    if not os.path.exists(TASKS_FILE):
        return {}
    with open(TASKS_FILE, 'r', encoding='utf-8') as file:
        return json.load(file)

def save_tasks(tasks):
    with open(TASKS_FILE, 'w', encoding='utf-8') as file:
        json.dump(tasks, file, indent=4, ensure_ascii=False)

@app.route('/save_task', methods=['POST'])
def save_task():
    team = session.get('team')  # 세션에서 팀 정보 가져오기
    task_date = request.form['date']
    task = request.form['task']

    if not task_date or not task or not team:
        return jsonify({'status': '잘못된 입력입니다.'}), 400

    tasks = load_tasks()

    if team not in tasks:
        tasks[team] = {}

    if task_date in tasks[team]:
        tasks[team][task_date].append(task)
    else:
        tasks[team][task_date] = [task]

    save_tasks(tasks)
    return jsonify({'status': '할 일 저장 완료!'})

@app.route('/get_tasks', methods=['GET'])
def get_tasks():
    team = session.get('team')  # 세션에서 팀 정보 가져오기
    task_date = request.args.get('date')

    if not task_date or not team:
        return jsonify([])

    tasks = load_tasks()
    team_tasks = tasks.get(team, {})
    return jsonify(team_tasks.get(task_date, []))

@app.route('/delete_task', methods=['POST'])
def delete_task():
    team = session.get('team')  # 세션에서 팀 정보 가져오기
    task_date = request.form['date']
    index = int(request.form['index'])

    tasks = load_tasks()

    if team in tasks and task_date in tasks[team] and 0 <= index < len(tasks[team][task_date]):
        tasks[team][task_date].pop(index)
        if not tasks[team][task_date]:  # 빈 리스트는 삭제
            del tasks[team][task_date]
        save_tasks(tasks)
        return jsonify({'status': '할 일 삭제 완료!'})
    return jsonify({'status': '할 일 삭제에 실패했습니다.'}), 400


def make_excel(team, task_month):
    tasks = load_tasks()
    wb = Workbook()
    ws = wb.active
    ws.title = task_month + "_task"
    path = "./excels/"

    ws.cell(row=1, column=1, value="날짜")
    ws.cell(row=1, column=2, value="내용")

    task_dic = tasks.get(team, {})
    sorted_task_dic = dict(sorted(task_dic.items()))
    print(sorted_task_dic)
    i = 0
    for key, value_list in sorted_task_dic.items():
        for index, value in enumerate(value_list):
            ws.cell(row=i+index+2, column=1, value=key)
            ws.cell(row=i+index+2, column=2, value=value)
        i += 1
    excel_name =  f"{path}{team} {ws.title}.xlsx"
    
    print(excel_name)
    wb.save(excel_name)
    return excel_name

@app.route('/calendar/download_excel')
def download_excel():
    team = session.get('team')
    task_month = request.args.get('task_month')
    
    if task_month is None:
        return "년월이 지정되지 않았습니다.", 400
    else:
        excel_name = make_excel(team, task_month)
        return send_file(f"{excel_name}", as_attachment=True)

@app.route('/weather', methods=['GET'])
def get_weather():
    date_str = request.args.get('date')  # 쿼리 문자열에서 'date' 파라미터 가져오기
    
    if not date_str:
        return jsonify({"error": "Date parameter is missing"}), 400

    # 요청된 날짜에 대한 날씨 정보 찾기
    weather_info = next((w for w in weather_data if w['date'] == date_str), {})
    return jsonify(weather_info)


def send_mail():
    send_email = os.getenv("SEND_EMAIL")
    send_pwd = os.getenv("SEND_PWD")
    smtp_name = "smtp.naver.com" 
    smtp_port = 587


    tasks = load_tasks()
    for user_id, value_dic in users.items():
        for key, value in value_dic.items():
            team = value_dic['team']
            recv_email = value_dic['email']
            
        if recv_email == None : continue   
        
        text = ""
        task_dic = tasks.get(team, {})
        

        for key, value_list in task_dic.items():
            if key == today:
                for value in value_list:
                    text += value + "\n"     

        msg = MIMEText(text, 'plain', 'utf-8') 
        msg['Subject'] = team + " " + today + "할 일 리스트입니다."  
        msg['From'] = send_email          
        msg['To'] = recv_email            

        email_string = msg.as_string()
        print(email_string)

        with smtplib.SMTP(smtp_name, smtp_port) as s:
            s.starttls()
            s.login(send_email, send_pwd)
            s.sendmail(send_email, recv_email, email_string)


    

#apscheduler 선언 
sched = BackgroundScheduler(daemon=True, timezone='Asia/Seoul') 


# 스케줄러가 이미 실행 중인지 확인
if not sched.get_jobs():
    # apscheduler실행설정, Cron방식으로, 00시 00분에 실행 
    sched.add_job(send_mail, 'cron', hour='00', minute='00') 
    # 스케줄링 실행
    sched.start()

if __name__ == '__main__':
    app.run(debug=True)